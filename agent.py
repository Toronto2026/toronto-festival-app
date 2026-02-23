# -*- coding: utf-8 -*-
"""
Агент автоматизации таблиц быстрого поиска
Фестиваль TORONTO (toronto.org.ua)

Використання:
    python agent.py --file "учасники_лютий_2026.xlsx" --type дипломи --date "28 лютого" --name "ТОРОНТО_ЛЮТИЙ_2026"
    python agent.py --file "учасники_лютий_2026.xlsx" --type подяки  --date "28 лютого" --name "ТОРОНТО_ЛЮТИЙ_2026"

Залежності:
    pip install pandas openpyxl

Для конвертації в PDF потрібен LibreOffice:
    Windows: https://www.libreoffice.org/download/download/
    Ubuntu:  sudo apt install libreoffice
"""

import argparse
import io
import math
import os
import subprocess
import sys

# Фикс кодировки — применяется только в main() при прямому запуску

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


def _log(*args):
    """Безпечний print — не падає якщо stdout закрито (напр. у Streamlit)."""
    try:
        print(*args)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Константи кольорів та шрифтів
# ---------------------------------------------------------------------------
COLOR_YELLOW  = "FFFF00"   # Artist / ПІБ керівника
COLOR_BLUE    = "ADD8E6"   # №Диплому / №Подяки
COLOR_GRAY    = "D3D3D3"   # решта колонок
COLOR_WHITE   = "FFFFFF"

FONT_NAME = "Arial"

# ---------------------------------------------------------------------------
# Ширини колонок (в символах / одиницях openpyxl)
# ---------------------------------------------------------------------------
COL_WIDTHS_DIPLOM = {
    "A": 6.44,   # ID
    "B": 23.0,   # Artist
    "C": 12.78,  # Номінація
    "D": 20.44,  # Назва або опис роботи
    "E": 9.55,   # Laureate
    "F": 10.66,  # №Диплому
}

COL_WIDTHS_PODYAKA = {
    "A": 8.0,    # ID
    "B": 40.0,   # ПІБ керівника
    "C": 12.0,   # №Подяки
}

# Приблизна кількість символів у рядку для кожної колонки
# (використовується для розрахунку висоти рядків)
CHARS_PER_ROW_DIPLOM = {
    "A": 9,
    "B": 16,
    "C": 13,
    "D": 22,
    "E": 11,
    "F": 8,
}

CHARS_PER_ROW_PODYAKA = {
    "A": 9,
    "B": 35,
    "C": 9,
}

# Висота одного текстового рядка у пунктах
LINE_HEIGHT = {
    16: 20,   # Arial 16pt
    11: 14,   # Arial 11pt
}

# ---------------------------------------------------------------------------
# Тексти опису (рядки 1–7)
# ---------------------------------------------------------------------------
DESC_DIPLOM = (
    "Вітаємо всіх учасників фестивалю з чудовими результатами!\n"
    "Як користуватися таблицею? Знаходимо у стовпчику Artist назву колективу або ПІБ учасника. "
    "Сортування за алфавітом. Навпроти ПІБ учасника ви бачите № диплома. "
    "У каталозі дипломів вибираємо свій № диплому та завантажуємо на свій гаджет.\n"
    "ВІДПРАВЛЕННЯ ПОСИЛОК З НАГОРОДАМИ ЗАПЛАНОВАНО НА {date}."
)

DESC_PODYAKA = (
    "Вітаємо всіх педагогів та учасників фестивалю з чудовими результатами!\n"
    "Як користуватися таблицею? Знаходимо у стовпчику ПІБ керівника. "
    "Сортування за алфавітом. Навпроти ПІБ керівника ви бачите № подяки. "
    "У каталозі подяк вибираємо свій № подяки та завантажуємо на свій гаджет.\n"
    "ВІДПРАВЛЕННЯ ПОСИЛОК З НАГОРОДАМИ ЗАПЛАНОВАНО НА {date}."
)

# ---------------------------------------------------------------------------
# Допоміжні функції
# ---------------------------------------------------------------------------

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def make_font(size: int, bold: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold)


def make_thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def make_center_alignment(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def make_left_alignment(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def calc_row_height(row_data: dict, chars_per_row: dict, font_sizes: dict) -> float:
    """
    Розраховує висоту рядка у пунктах.

    row_data      – {col_letter: text}
    chars_per_row – {col_letter: кількість символів у рядку}
    font_sizes    – {col_letter: розмір шрифту}
    """
    max_lines = 1
    for col, text in row_data.items():
        if not isinstance(text, str):
            text = "" if text is None else str(text)
        paragraphs = text.split("\n")
        total_lines = 0
        width = chars_per_row.get(col, 10)
        for para in paragraphs:
            total_lines += math.ceil(len(para) / width) if width > 0 else 1
            total_lines = max(total_lines, 1)
        max_lines = max(max_lines, total_lines)

    # Беремо шрифт першої не-ID колонки (зазвичай 16pt для Artist)
    # Але для точності беремо максимальну висоту серед усіх колонок
    max_height = 0.0
    for col in row_data:
        n_lines = 1
        text = row_data[col]
        if not isinstance(text, str):
            text = "" if text is None else str(text)
        paragraphs = text.split("\n")
        n_lines = 0
        width = chars_per_row.get(col, 10)
        for para in paragraphs:
            n_lines += math.ceil(len(para) / width) if width > 0 else 1
            n_lines = max(n_lines, 1)
        fsize = font_sizes.get(col, 11)
        lh = LINE_HEIGHT.get(fsize, 14)
        cell_height = n_lines * lh + 8
        max_height = max(max_height, cell_height)

    return max(max_height, 30.0)


def set_print_settings(ws, last_col_letter: str, header_row: int = 8) -> None:
    """Налаштовує параметри друку: А4, книжна, fit to width, повтор заголовків."""
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage   = True

    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    ws.print_title_rows = f"{header_row}:{header_row}"


def sort_key(val):
    """Ключ сортування: None/порожній рядок — в кінець."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return (1, "")
    return (0, str(val).strip().lower())


def convert_to_pdf(xlsx_path: str) -> bool:
    """
    Конвертує xlsx у PDF через LibreOffice headless.
    Повертає True при успіху.
    """
    # Шукаємо LibreOffice
    candidates = [
        "libreoffice",
        "soffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/libreoffice",
        "/usr/bin/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]

    lo_bin = None
    for c in candidates:
        if os.path.isfile(c):
            lo_bin = c
            break
        # Спроба через PATH
        try:
            result = subprocess.run(
                [c, "--version"],
                capture_output=True, timeout=10
            )
            if result.returncode == 0:
                lo_bin = c
                break
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue

    if lo_bin is None:
        _log("[!]  LibreOffice не знайдено. PDF не створено.")
        _log("   Встановіть LibreOffice: https://www.libreoffice.org/download/download/")
        return False

    out_dir = os.path.dirname(os.path.abspath(xlsx_path))
    cmd = [
        lo_bin, "--headless", "--convert-to", "pdf",
        "--outdir", out_dir,
        os.path.abspath(xlsx_path)
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode == 0:
            _log(f"[OK] PDF створено: {os.path.splitext(xlsx_path)[0]}.pdf")
            return True
        else:
            _log(f"[ERROR] Помилка LibreOffice: {result.stderr}")
            return False
    except subprocess.TimeoutExpired:
        _log("[ERROR] Timeout: LibreOffice не відповідає.")
        return False


# ---------------------------------------------------------------------------
# Генератор таблиці "Дипломи онлайн"
# ---------------------------------------------------------------------------

def build_diplom_table(df: pd.DataFrame, date: str, out_name: str) -> str:
    """
    Формує xlsx-файл таблиці дипломів.
    Повертає шлях до збереженого файлу.
    """
    # --- Знаходимо потрібні колонки (регістронезалежно) ---
    col_map = {}
    for col in df.columns:
        cl = col.strip().lower()
        if cl == "id":
            col_map["ID"] = col
        elif ("піб учасника" in cl or "artist" in cl or "pib" in cl
              or ("піб" in cl and "керівник" not in cl and "концертмейстер" not in cl)):
            col_map["Artist"] = col
        elif "номінац" in cl or "nomination" in cl:
            col_map["Номінація"] = col
        elif ("назва" in cl or "опис" in cl or "description" in cl or "title" in cl):
            col_map["Назва"] = col
        elif "laureate" in cl or "лауреат" in cl:
            col_map["Laureate"] = col
        elif "диплом" in cl or "diploma" in cl:
            col_map["№Диплому"] = col

    # №Диплому — обов'язкова колонка. Якщо її немає — попереджаємо, але не зупиняємось:
    # колонка буде порожньою (оператор заповнює вручну).
    missing_required = [k for k in ("ID", "Artist") if k not in col_map]
    if missing_required:
        _log(f"[ERROR] Не знайдено обов'язкових колонок: {missing_required}")
        _log(f"   Наявні колонки: {list(df.columns)}")
        sys.exit(1)

    auto_numbered = "№Диплому" not in col_map
    if auto_numbered:
        _log("[i]  Колонка '№ Диплома' не знайдена — нумерується за порядком вихідного файлу, потім сортування.")

    # --- Формуємо робочий DataFrame ---
    cols_order = ["ID", "Artist", "Номінація", "Назва", "Laureate", "№Диплому"]
    work_df = pd.DataFrame()
    for key in cols_order:
        if key in col_map:
            work_df[key] = df[col_map[key]].fillna("").astype(str).str.strip()
        else:
            work_df[key] = ""

    # --- Автонумерація дипломів (ДО сортування — за порядком вихідного файлу) ---
    if auto_numbered:
        work_df["№Диплому"] = [str(i) for i in range(1, len(work_df) + 1)]

    # --- Сортування за Artist ---
    work_df = work_df.sort_values(
        by="Artist",
        key=lambda s: s.map(sort_key),
        kind="stable"
    ).reset_index(drop=True)

    # --- Створюємо Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Дипломи онлайн"

    # Ширини колонок
    for col_letter, width in COL_WIDTHS_DIPLOM.items():
        ws.column_dimensions[col_letter].width = width

    # --- Рядки 1–7: блок опису ---
    ws.merge_cells("A1:F7")
    desc_cell = ws["A1"]
    desc_cell.value = DESC_DIPLOM.format(date=date)
    desc_cell.fill      = make_fill(COLOR_YELLOW)
    desc_cell.font      = make_font(16, bold=True)
    desc_cell.alignment = make_center_alignment(wrap=True)
    ws.row_dimensions[1].height = 130  # 7 рядків об'єднані — фіксована висота блоку

    # --- Рядок 8: заголовки ---
    headers = ["ID", "Artist", "Номінація", "Назва або опис роботи", "Laureate", "№Диплому"]
    header_fills = [
        make_fill(COLOR_GRAY),    # ID
        make_fill(COLOR_YELLOW),  # Artist
        make_fill(COLOR_GRAY),    # Номінація
        make_fill(COLOR_GRAY),    # Назва
        make_fill(COLOR_GRAY),    # Laureate
        make_fill(COLOR_BLUE),    # №Диплому
    ]
    header_font_sizes = [11, 16, 11, 11, 11, 16]

    for col_idx, (header, fill, fsize) in enumerate(
            zip(headers, header_fills, header_font_sizes), start=1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.fill      = fill
        cell.font      = make_font(fsize, bold=True)
        cell.alignment = make_center_alignment()
        cell.border    = make_thin_border()
    ws.row_dimensions[8].height = 30

    # --- Рядки 9+: дані ---
    data_fills = [
        make_fill(COLOR_GRAY),    # ID
        make_fill(COLOR_YELLOW),  # Artist
        make_fill(COLOR_GRAY),    # Номінація
        make_fill(COLOR_GRAY),    # Назва
        make_fill(COLOR_GRAY),    # Laureate
        make_fill(COLOR_BLUE),    # №Диплому
    ]
    data_fonts = [
        make_font(11),
        make_font(16),
        make_font(11),
        make_font(11),
        make_font(11),
        make_font(16, bold=True),
    ]
    data_alignments = [
        make_left_alignment(wrap=False),   # ID
        make_left_alignment(wrap=True),    # Artist
        make_left_alignment(wrap=True),    # Номінація
        make_left_alignment(wrap=True),    # Назва
        make_left_alignment(wrap=False),   # Laureate
        make_center_alignment(wrap=False), # №Диплому
    ]
    col_letters = ["A", "B", "C", "D", "E", "F"]
    font_sizes_map = {"A": 11, "B": 16, "C": 11, "D": 11, "E": 11, "F": 16}

    for row_idx, (_, row) in enumerate(work_df.iterrows(), start=9):
        values = [
            row["ID"],
            row["Artist"],
            row["Номінація"],
            row["Назва"],
            row["Laureate"],
            row["№Диплому"],
        ]
        row_data_for_height = {}
        for col_idx, (val, fill, font, align, col_letter) in enumerate(
                zip(values, data_fills, data_fonts, data_alignments, col_letters),
                start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = align
            cell.border    = make_thin_border()
            row_data_for_height[col_letter] = str(val) if val is not None else ""

        # Висота рядка
        h = calc_row_height(row_data_for_height, CHARS_PER_ROW_DIPLOM, font_sizes_map)
        ws.row_dimensions[row_idx].height = h

    # --- Параметри друку ---
    set_print_settings(ws, last_col_letter="F", header_row=8)

    # --- Зберігаємо ---
    # out_name може бути абсолютним шляхом (з app.py) або просто назвою (з CLI)
    out_dir  = os.path.dirname(out_name) if os.path.dirname(out_name) else "."
    out_base = os.path.basename(out_name)
    out_path = os.path.join(out_dir, f"ДИПЛОМ_ОНЛАЙН_{out_base}.xlsx")
    wb.save(out_path)
    _log(f"[OK] Збережено: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# Генератор таблиці "Подяки онлайн"
# ---------------------------------------------------------------------------

def build_podyaka_table(df: pd.DataFrame, date: str, out_name: str) -> str:
    """
    Формує xlsx-файл таблиці подяк.
    Повертає шлях до збереженого файлу.
    """
    # --- Знаходимо потрібні колонки ---
    col_map = {}
    for col in df.columns:
        cl = col.strip().lower()
        if cl == "id":
            col_map["ID"] = col
        elif ("керівник" in cl or "концертмейстер" in cl
              or "пед" in cl or "teacher" in cl
              or ("піб" in cl and "учасник" not in cl)):
            col_map["ПІБ"] = col
        elif "подяк" in cl or "gratitude" in cl or "подяка" in cl:
            col_map["№Подяки"] = col

    # Якщо не знайшли ПІБ — беремо другу колонку після ID
    if "ПІБ" not in col_map:
        non_id_cols = [c for c in df.columns if c.strip().lower() != "id"]
        if non_id_cols:
            col_map["ПІБ"] = non_id_cols[0]
            _log(f"[i]  Колонка ПІБ керівника: '{non_id_cols[0]}'")

    if "ID" not in col_map or "ПІБ" not in col_map:
        _log(f"[ERROR] Не знайдено обов'язкових колонок для подяк.")
        _log(f"   Наявні колонки: {list(df.columns)}")
        sys.exit(1)

    # --- Формуємо робочий DataFrame ---
    work_df = pd.DataFrame()
    work_df["ID"]  = df[col_map["ID"]].fillna("").astype(str).str.strip()
    work_df["ПІБ"] = df[col_map["ПІБ"]].fillna("").astype(str).str.strip()

    # №Подяки — якщо є в файлі, беремо; якщо ні — нумеруємо по порядку до сортування
    if "№Подяки" in col_map:
        work_df["№Подяки"] = df[col_map["№Подяки"]].fillna("").astype(str).str.strip()
    else:
        work_df["№Подяки"] = [str(i) for i in range(1, len(work_df) + 1)]

    # --- Сортування за ПІБ ---
    work_df = work_df.sort_values(
        by="ПІБ",
        key=lambda s: s.map(sort_key),
        kind="stable"
    ).reset_index(drop=True)

    # --- Створюємо Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Подяки онлайн"

    # Ширини колонок
    for col_letter, width in COL_WIDTHS_PODYAKA.items():
        ws.column_dimensions[col_letter].width = width

    # --- Рядки 1–7: блок опису ---
    ws.merge_cells("A1:C7")
    desc_cell = ws["A1"]
    desc_cell.value = DESC_PODYAKA.format(date=date)
    desc_cell.fill      = make_fill(COLOR_YELLOW)
    desc_cell.font      = make_font(16, bold=True)
    desc_cell.alignment = make_center_alignment(wrap=True)
    ws.row_dimensions[1].height = 130

    # --- Рядок 8: заголовки ---
    headers = ["ID", "ПІБ керівника, концертмейстера", "№Подяки"]
    header_fills = [
        make_fill(COLOR_GRAY),
        make_fill(COLOR_YELLOW),
        make_fill(COLOR_BLUE),
    ]
    header_font_sizes = [11, 16, 16]

    for col_idx, (header, fill, fsize) in enumerate(
            zip(headers, header_fills, header_font_sizes), start=1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.fill      = fill
        cell.font      = make_font(fsize, bold=True)
        cell.alignment = make_center_alignment()
        cell.border    = make_thin_border()
    ws.row_dimensions[8].height = 30

    # --- Рядки 9+: дані ---
    data_fills = [
        make_fill(COLOR_GRAY),
        make_fill(COLOR_YELLOW),
        make_fill(COLOR_BLUE),
    ]
    data_fonts = [
        make_font(11),
        make_font(16),
        make_font(16, bold=True),
    ]
    data_alignments = [
        make_left_alignment(wrap=False),
        make_left_alignment(wrap=True),
        make_center_alignment(wrap=False),
    ]
    col_letters = ["A", "B", "C"]
    font_sizes_map = {"A": 11, "B": 16, "C": 16}

    for row_idx, (_, row) in enumerate(work_df.iterrows(), start=9):
        values = [row["ID"], row["ПІБ"], row["№Подяки"]]
        row_data_for_height = {}
        for col_idx, (val, fill, font, align, col_letter) in enumerate(
                zip(values, data_fills, data_fonts, data_alignments, col_letters),
                start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = align
            cell.border    = make_thin_border()
            row_data_for_height[col_letter] = str(val) if val is not None else ""

        # Висота рядка
        h = calc_row_height(row_data_for_height, CHARS_PER_ROW_PODYAKA, font_sizes_map)
        ws.row_dimensions[row_idx].height = h

    # --- Параметри друку ---
    set_print_settings(ws, last_col_letter="C", header_row=8)

    # --- Зберігаємо ---
    out_dir  = os.path.dirname(out_name) if os.path.dirname(out_name) else "."
    out_base = os.path.basename(out_name)
    out_path = os.path.join(out_dir, f"ПОДЯКИ_ОНЛАЙН_{out_base}.xlsx")
    wb.save(out_path)
    _log(f"[OK] Збережено: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# Точка входу
# ---------------------------------------------------------------------------

def main():
    # Фикс кодировки консоли Windows (только при прямом запуске, не через Streamlit)
    _in_streamlit = "streamlit" in sys.modules
    if sys.platform == "win32" and not _in_streamlit and hasattr(sys.stdout, "buffer"):
        try:
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
        except Exception:
            pass

    parser = argparse.ArgumentParser(
        description="Агент формування таблиць швидкого пошуку — Фестиваль TORONTO"
    )
    parser.add_argument(
        "--file", required=True,
        help="Шлях до вхідного xlsx-файлу учасників"
    )
    parser.add_argument(
        "--type", required=True, choices=["дипломи", "подяки"],
        help="Тип таблиці: 'дипломи' або 'подяки'"
    )
    parser.add_argument(
        "--date", required=True,
        help="Дата відправлення нагород (напр. 'Навесні 2026' або '28 лютого')"
    )
    parser.add_argument(
        "--name", required=True,
        help="Назва для вихідних файлів (напр. 'ТОРОНТО_ЛЮТИЙ_2026')"
    )
    parser.add_argument(
        "--no-pdf", action="store_true",
        help="Не конвертувати в PDF (лише xlsx)"
    )

    args = parser.parse_args()

    # Перевірка вхідного файлу
    if not os.path.isfile(args.file):
        print(f"[ERROR] Файл не знайдено: {args.file}")
        sys.exit(1)

    print(f"[-&gt;] Читаю файл: {args.file}")
    try:
        df = pd.read_excel(args.file, engine="openpyxl")
    except Exception as e:
        print(f"[ERROR] Помилка читання файлу: {e}")
        sys.exit(1)

    print(f"   Рядків: {len(df)}, Колонок: {list(df.columns)}")

    # Генерація таблиці
    table_type = args.type.lower()
    if table_type == "дипломи":
        xlsx_path = build_diplom_table(df, date=args.date, out_name=args.name)
    else:
        xlsx_path = build_podyaka_table(df, date=args.date, out_name=args.name)

    # Конвертація в PDF
    if not args.no_pdf:
        print("[..] Конвертую в PDF...")
        convert_to_pdf(xlsx_path)
    else:
        print("[i]  PDF пропущено (--no-pdf)")

    print("\n[OK] Готово!")


if __name__ == "__main__":
    main()
